import ccxt
import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def format_volume(vol):
    if vol >= 1_000_000:
        return f"{vol / 1_000_000:.2f}M"
    elif vol >= 1_000:
        return f"{vol / 1_000:.2f}K"
    else:
        return str(round(vol, 2))

def calculate_qqe(df, rsi_length=14, smoothing_factor=5):
    delta = df['close'].diff()
    gain = np.where(delta > 0, delta, 0)
    loss = np.where(delta < 0, -delta, 0)
    avg_gain = pd.Series(gain).rolling(rsi_length).mean()
    avg_loss = pd.Series(loss).rolling(rsi_length).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    rsi = rsi.bfill()

    rsii = rsi.ewm(span=smoothing_factor, adjust=False).mean()
    tr = abs(rsii - rsii.shift(1))
    wwalpha = 1 / rsi_length
    wwma = tr.ewm(alpha=wwalpha, adjust=False).mean()
    atrrsi = wwma.ewm(alpha=wwalpha, adjust=False).mean()
    qqef = rsii
    qup = qqef + 4.236 * atrrsi
    qdn = qqef - 4.236 * atrrsi

    qqes = pd.Series(np.zeros(len(df)))
    for i in range(1, len(df)):
        if qup[i] < qqes[i-1]:
            qqes[i] = qup[i]
        elif qqef[i] > qqes[i-1] and qqef[i-1] < qqes[i-1]:
            qqes[i] = qdn[i]
        elif qdn[i] > qqes[i-1]:
            qqes[i] = qdn[i]
        elif qqef[i] < qqes[i-1] and qqef[i-1] > qqes[i-1]:
            qqes[i] = qup[i]
        else:
            qqes[i] = qqes[i-1]
    return qqef, qqes

def check_signals(qqef, qqes):
    signal = "NO SIGNAL"
    for i in range(-2, 0):
        if qqef.iloc[i-1] < qqes.iloc[i-1] and qqef.iloc[i] > qqes.iloc[i]:
            return "BUY"
        elif qqef.iloc[i-1] > qqes.iloc[i-1] and qqef.iloc[i] < qqes.iloc[i]:
            return "SELL"
    return signal

def main():
    print("QQE sinyal taraması başladı...")
    exchange = ccxt.binance()
    timeframe = '4h'
    limit = 100

    markets = exchange.load_markets()
    symbols = [s for s in markets if s.endswith('/USDT') and len(s.split('/')) == 2]

    results = []

    for idx, symbol in enumerate(symbols, 1):
        print(f"{idx}. {symbol} taranıyor...")
        try:
            ohlcv = exchange.fetch_ohlcv(symbol, timeframe=timeframe, limit=limit)
            df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
            df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')

            df['usdt_volume'] = df['close'] * df['volume']
            qqef, qqes = calculate_qqe(df)
            signal = check_signals(qqef, qqes)

            raw_volume = df['usdt_volume'].iloc[-6:].sum()
            volume_24h = format_volume(raw_volume)

            if signal in ['BUY', 'SELL']:
                results.append({
                    "Sembol": symbol,
                    "Sinyal": signal,
                    "24h Hacim (USDT)": volume_24h
                })

            time.sleep(0.2)

        except Exception as e:
            print(f"{symbol}: Hata - {e}")

    # Excel yazımı
    if results:
        df_results = pd.DataFrame(results)
        now = datetime.now().strftime("%d.%m.%Y - %H.%M")
        filename = f"qqe_sinyal_{now}.xlsx"
        df_results.to_excel(filename, index=False)

        wb = load_workbook(filename)
        ws = wb.active

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            signal_cell = ws[f"B{row}"]
            if signal_cell.value == "BUY":
                for col in ['A', 'B', 'C']:
                    ws[f"{col}{row}"].fill = green_fill
            elif signal_cell.value == "SELL":
                for col in ['A', 'B', 'C']:
                    ws[f"{col}{row}"].fill = red_fill

        wb.save(filename)
        print(f"\nSinyaller başarıyla kaydedildi: {filename}")
    else:
        print("Sinyal bulunamadı.")

if __name__ == "__main__":
    main()
